"""
Utility functions for model metrics and evaluation.

This module provides functions to calculate regression metrics, process model predictions,
and generate various plots for uncertainty quantification and calibration analysis.

"""

import logging
import math
import os
from bisect import bisect_left
from pathlib import Path
from typing import Tuple, Dict, Union, Optional, List, Any, Callable

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import scipy.stats as ss
import seaborn as sns
import torch
import uncertainty_toolbox as uct
import wandb
from matplotlib.ticker import MaxNLocator
from openpyxl import load_workbook
from scipy.integrate import quad
from scipy.stats import bootstrap
from scipy.stats import norm
from scipy.stats import spearmanr
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score, mean_squared_error, explained_variance_score

from uncertainty_toolbox import viz as uct_viz
from uncertainty_toolbox.metrics import get_all_metrics, METRIC_NAMES
from wandb import Image

from uqdd import FIGS_DIR, DATA_DIR
from uqdd.utils import create_logger, save_df, save_pickle

string_types = (type(b""), type(""))
sns.set_theme(style="white")


def _rmse(a: np.ndarray, b: np.ndarray) -> float:
    """Compute RMSE without relying on sklearn's squared kwarg (compatibility)."""
    return float(np.sqrt(np.mean((a - b) ** 2)))


def calc_nanaware_metrics(
        tensor: torch.Tensor,
        nan_mask: torch.Tensor,
        all_tasks_agg: Union[bool, str] = False,
) -> torch.Tensor:
    """
    Calculate metrics while handling NaN values in an input tensor.

    Parameters
    ----------
    tensor : torch.Tensor
        Input tensor containing numerical values.
    nan_mask : torch.Tensor
        Boolean mask (same shape as ``tensor``) indicating NaN positions.
    all_tasks_agg : bool or {"mean", "sum"}, optional
        Aggregation behavior across tasks: ``False`` returns per-task means;
        ``"mean"`` returns mean across tasks; ``"sum"`` returns sum across tasks.

    Returns
    -------
    torch.Tensor
        Aggregated tensor according to ``all_tasks_agg``.
    """
    valid_values = torch.where(
        ~nan_mask, tensor, torch.tensor(0.0, device=tensor.device)
    )
    sum_values = torch.sum(valid_values, dim=0)
    valid_counts = torch.sum(~nan_mask, dim=0)

    tensor_means = sum_values / valid_counts.clamp(min=1)

    if all_tasks_agg == "mean":
        return torch.nanmean(tensor_means)
    elif all_tasks_agg == "sum":
        return torch.nansum(tensor_means)
    return tensor_means


def calc_regr_metrics(
        targets: torch.Tensor, outputs: torch.Tensor, metrics_per_task: bool = False
) -> Tuple[float, float, float]:
    """
    Compute regression metrics: RMSE, R^2, and explained variance.

    Parameters
    ----------
    targets : torch.Tensor
        True target values.
    outputs : torch.Tensor
        Model-predicted values.
    metrics_per_task : bool, optional
        If True, compute metrics per task (in multitask). Default is ``False``.

    Returns
    -------
    (float, float, float)
        RMSE, R^2, and explained variance scores.
    """
    if outputs.dim() > targets.dim():
        outputs = outputs.nanmean(dim=-1)

    targets = targets.detach().cpu()
    outputs = outputs.detach().cpu()

    if metrics_per_task and targets.ndim == 2 and targets.shape[1] > 1:
        rmse = np.full(targets.shape[1], np.nan, dtype=float)
        r2 = np.full(targets.shape[1], np.nan, dtype=float)
        evs = np.full(targets.shape[1], np.nan, dtype=float)

        for i in range(targets.shape[1]):
            task_t = targets[:, i]
            task_o = outputs[:, i]
            valid_mask = ~torch.isnan(task_t)
            if valid_mask.any():
                task_t = task_t[valid_mask].numpy().astype(float)
                task_o = task_o[valid_mask].numpy().astype(float)
                rmse[i] = _rmse(task_t, task_o)
                r2[i] = r2_score(task_t, task_o)
                evs[i] = explained_variance_score(task_t, task_o)
        # Return arrays cast to floats if single-valued
        return float(np.nanmean(rmse)), float(np.nanmean(r2)), float(np.nanmean(evs))
    else:
        nan_mask = ~torch.isnan(targets)
        t = targets[nan_mask].numpy().astype(float).flatten()
        o = outputs[nan_mask].numpy().astype(float).flatten()
        rmse = _rmse(t, o)
        r2 = r2_score(t, o)
        evs = explained_variance_score(t, o)

    return float(rmse), float(r2), float(evs)


def process_preds(
        predictions: Union[torch.Tensor, np.ndarray],
        targets: Union[torch.Tensor, np.ndarray],
        alea_vars: Optional[Union[torch.Tensor, np.ndarray]] = None,
        epi_vars: Optional[Union[torch.Tensor, np.ndarray]] = None,
        task_idx: Optional[int] = None,
        model_type: Optional[str] = None,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    """
    Process predictions to extract mean/variance, align with targets, and compute errors.

    Parameters
    ----------
    predictions : torch.Tensor or numpy.ndarray
        Model predictions with dims [samples, tasks, ensemble_members] (for ensembles).
    targets : torch.Tensor or numpy.ndarray
        True target values.
    alea_vars : torch.Tensor or numpy.ndarray or None, optional
        Aleatoric uncertainty estimates.
    epi_vars : torch.Tensor or numpy.ndarray or None, optional
        Epistemic uncertainty estimates.
    task_idx : int or None, optional
        Specific task index to select in multitask.
    model_type : str or None, optional
        Type of model (e.g., "ensemble", "mcdropout", "evidential", "pnn").

    Returns
    -------
    (np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray)
        y_true, y_pred, y_err, y_alea, y_eps.
    """
    if model_type == "pnn":
        y_pred = predictions
        y_alea = alea_vars
        y_eps = torch.zeros_like(alea_vars)
    elif model_type in ["ensemble", "mcdropout"]:
        y_pred = predictions.mean(dim=-1)
        y_alea = alea_vars.mean(dim=-1)
        y_eps = predictions.var(dim=-1)
    elif model_type in ["evidential", "eoe", "emc", "pnn"]:
        y_pred = predictions
        y_alea = alea_vars
        y_eps = epi_vars
    else:
        raise ValueError(f"Unknown model type: {model_type}")

    y_true = targets

    if task_idx is not None:
        y_true, y_pred, y_alea, y_eps = (
            y_true[:, task_idx],
            y_pred[:, task_idx],
            y_alea[:, task_idx],
            y_eps[:, task_idx],
        )

    nan_mask = ~torch.isnan(y_true)
    y_true, y_pred, y_alea, y_eps = (
        y_true[nan_mask],
        y_pred[nan_mask],
        y_alea[nan_mask],
        y_eps[nan_mask],
    )

    y_err = y_pred - y_true

    y_true, y_pred, y_err, y_alea, y_eps = map(
        lambda x: x.cpu().numpy(), (y_true, y_pred, y_err, y_alea, y_eps)
    )

    return y_true, y_pred, y_err, y_alea, y_eps


def get_preds_export_path(data_specific_path: str, model_name: str) -> Path:
    """
    Build the file path for exporting prediction results.

    Parameters
    ----------
    data_specific_path : str
        Dataset-specific directory path under ``DATA_DIR/predictions``.
    model_name : str
        Trained model name.

    Returns
    -------
    Path
        Path where the predictions CSV file will be saved.
    """
    path = DATA_DIR / "predictions" / Path(data_specific_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    return path / f"{model_name}_preds.csv"


def create_df_preds(
        y_true: np.ndarray,
        y_pred: np.ndarray,
        y_err: np.ndarray,
        y_alea: Optional[np.ndarray] = None,
        y_eps: Optional[np.ndarray] = None,
        export: bool = True,
        data_specific_path: Optional[str | Path] = None,
        model_name: Optional[str] = None,
        logger: Optional[logging.Logger] = None,
) -> pd.DataFrame:
    """
    Create a DataFrame containing predictions, uncertainties, and errors.

    Parameters
    ----------
    y_true : numpy.ndarray
        Ground-truth target values.
    y_pred : numpy.ndarray
        Predicted values.
    y_err : numpy.ndarray
        Prediction errors (y_pred - y_true).
    y_alea : numpy.ndarray or None, optional
        Aleatoric uncertainty estimates.
    y_eps : numpy.ndarray or None, optional
        Epistemic uncertainty estimates.
    export : bool, optional
        If True, export the DataFrame as CSV. Default is ``True``.
    data_specific_path : str or Path or None, optional
        Path for saving the CSV file.
    model_name : str or None, optional
        Name of the model.
    logger : logging.Logger or None, optional
        Logger instance.

    Returns
    -------
    pd.DataFrame
        DataFrame containing predictions and uncertainty estimates.
    """
    df = pd.DataFrame(
        {
            "y_true": y_true,
            "y_pred": y_pred,
            "y_err": y_err,
            "y_alea": y_alea,
            "y_eps": y_eps,
        }
    )

    if export and data_specific_path and model_name:
        export_path = get_preds_export_path(data_specific_path, model_name)
        save_df(df, str(export_path))
        if logger:
            logger.debug(f"Exported predictions to {export_path}")

    return df


def plot_true_vs_preds(
        y_pred: np.ndarray,
        y_true: np.ndarray,
        n_subset: Optional[int] = None,
        distant_threshold: float = 3.0,
        savefig: bool = False,
        save_dir: Optional[Path] = None,
        task_name: Optional[str] = None,
) -> plt.Figure:
    """
    Generate a scatter plot comparing true vs. predicted values.

    Parameters
    ----------
    y_pred : np.ndarray
        Predicted values.
    y_true : np.ndarray
        True values.
    n_subset : Optional[int], optional
        Number of samples to plot, by default None.
    distant_threshold : float, optional
        Threshold to identify distant points, by default 3.0.
    savefig : bool, optional
        Whether to save the figure, by default False.
    save_dir : Optional[Path], optional
        Directory to save the figure, by default None.
    task_name : Optional[str], optional
        Name of the task, by default None.

    Returns
    -------
    plt.Figure
        The generated scatter plot figure.
    """
    if n_subset is not None and n_subset < len(y_true):
        rng = np.random.default_rng(42)
        indices = rng.choice(len(y_true), size=n_subset, replace=False)

        y_true_subset = y_true[indices]
        y_preds_subset = y_pred[indices]
    else:
        y_true_subset = y_true
        y_preds_subset = y_pred

    # Create a DataFrame for seaborn
    data = pd.DataFrame(
        {"True Values": y_true_subset, "Predicted Values": y_preds_subset}
    )

    # Calculate RMSE for the subset
    rmse = np.sqrt(np.mean((data["Predicted Values"] - data["True Values"]) ** 2))

    # Identify distant/erroneous points
    data["Distant"] = (
            np.abs(data["True Values"] - data["Predicted Values"])
            > distant_threshold * rmse
    )
    # Calculate the errors
    data["Error"] = np.abs(data["True Values"] - data["Predicted Values"])

    g = sns.JointGrid(data=data, x="True Values", y="Predicted Values")

    # Add the marginal histograms without hue
    sns.histplot(
        data=data,
        x="True Values",
        ax=g.ax_marg_x,
        bins=50,
        fill=True,
        color="gray",
        kde=True,
    )  # green
    sns.histplot(
        data=data,
        y="Predicted Values",
        ax=g.ax_marg_y,
        bins=50,
        fill=True,
        color="gray",
        kde=True,
    )  # 'green'

    # Set identical integer ticks for x and y axes
    g.ax_joint.xaxis.set_major_locator(MaxNLocator(integer=True))
    g.ax_joint.yaxis.set_major_locator(MaxNLocator(integer=True))
    # Add identity line
    min_val = min(g.ax_joint.get_xlim()[0], g.ax_joint.get_ylim()[0])
    max_val = max(g.ax_joint.get_xlim()[1], g.ax_joint.get_ylim()[1])
    g.ax_joint.set_xlim(min_val, max_val)
    g.ax_joint.set_ylim(min_val, max_val)
    g.ax_joint.plot(
        [min_val, max_val],
        [min_val, max_val],
        linestyle="--",
        color="gray",
        alpha=0.5,
        label=None,
    )
    g.ax_joint.fill_between(
        np.arange(int(min_val), int(max_val) + 1),
        np.arange(int(min_val), int(max_val) + 1) - rmse,
        np.arange(int(min_val), int(max_val) + 1) + rmse,
        color="mediumblue",
        alpha=0.2,
        label=f"±1 RMSE ({rmse:.2f})",
    )
    g.ax_joint.fill_between(
        np.arange(int(min_val), int(max_val) + 1),
        np.arange(int(min_val), int(max_val) + 1) - 2 * rmse,
        np.arange(int(min_val), int(max_val) + 1) + 2 * rmse,
        color="skyblue",
        alpha=0.5,
        label=f"±2 RMSE ({2 * rmse:.2f})",
    )

    # Add the scatter plot with colors
    sns.scatterplot(
        data=data,
        x="True Values",
        y="Predicted Values",
        ax=g.ax_joint,
        size=20,
        alpha=0.5,
        hue="Distant",
        palette={
            True: "darkorange",
            False: "cornflowerblue",
        },  # indianred, cornflowerblue
    )
    sns.regplot(
        data=data,
        x="True Values",
        y="Predicted Values",
        ax=g.ax_joint,
        scatter=False,
        ci=95,
        n_boot=1000,
    )

    within_1_rmse = (
            np.mean((np.abs(data["True Values"] - data["Predicted Values"]) <= rmse)) * 100
    )
    within_2_rmse = (
            np.mean((np.abs(data["True Values"] - data["Predicted Values"]) <= 2 * rmse))
            * 100
    )
    distant_percentage = np.mean(data["Distant"]) * 100

    legend_elements = [
        plt.Line2D(
            [0],
            [0],
            color="mediumblue",
            lw=4,
            alpha=0.2,
            label=f"±1 RMSE ({rmse:.2f}): {within_1_rmse:.2f}%",
        ),
        plt.Line2D(
            [0],
            [0],
            color="skyblue",
            lw=4,
            alpha=0.5,
            label=f"±2 RMSE ({2 * rmse:.2f}): {within_2_rmse:.2f}%",
        ),
        plt.Line2D(
            [0],
            [0],
            color="darkorange",
            alpha=0.5,
            marker="o",
            linestyle="",
            markersize=8,
            label=f"> {distant_threshold} RMSE: {distant_percentage:.2f}%",
        ),  # indianred
    ]
    legend = g.ax_joint.legend(
        handles=legend_elements,
        title="RMSE Bands",
        loc="upper left",
        # labels=legend_text,
        fontsize=8,
        title_fontsize=10,
        handletextpad=0.5,
        labelspacing=0.5,
        borderaxespad=0.5,
    )

    if savefig:
        g.savefig(
            save_dir / f"true_vs_preds_plot_with_errs_{task_name}.png",
            format="png",
            bbox_inches="tight",
            dpi=1200,
        )
        g.savefig(
            save_dir / f"true_vs_preds_plot_with_errs_{task_name}.svg",
            format="svg",
            bbox_inches="tight",
        )
        g.savefig(
            save_dir / f"true_vs_preds_plot_with_errs_{task_name}.pdf",
            format="pdf",
            bbox_inches="tight",
        )
        g.savefig(
            save_dir / f"true_vs_preds_plot_with_errs_{task_name}.eps",
            format="eps",
            bbox_inches="tight",
        )
    return g.figure


def plot_pred_intervals(
        y_pred: np.ndarray,
        y_std: np.ndarray,
        y_true: np.ndarray,
        n_subset: int = 100,
        ylims: Optional[Tuple[float, float]] = (-3.0, 3.0),
        num_stds_confidence_bound: float = 2.0,
        ax: Optional[plt.Axes] = None,
        ordered: bool = False,
) -> None:
    """
    Plots prediction intervals to visualize uncertainty.

    Parameters
    ----------
    y_pred : np.ndarray
        Predicted values.
    y_std : np.ndarray
        Standard deviation / Uncertainty of predictions.
    y_true : np.ndarray
        True values.
    n_subset : int, optional
        Number of samples to plot, by default 100.
    ylims : Optional[Tuple[float, float]], optional
        Y-axis limits, by default (-3.0, 3.0).
    num_stds_confidence_bound : float, optional
        Number of standard deviations for confidence bounds, by default 2.0.
    ax : Optional[plt.Axes], optional
        Matplotlib axis to plot on, by default None.
    ordered : bool, optional
        Whether to order the intervals, by default False.

    Returns
    -------
    None
    """
    if ax is None:
        fig, ax = plt.subplots()
    if ordered:
        uct_viz.plot_intervals_ordered(
            y_pred=y_pred,
            y_std=y_std,
            y_true=y_true,
            n_subset=n_subset,
            ylims=ylims,
            num_stds_confidence_bound=int(num_stds_confidence_bound),
            ax=ax,
        )
    else:
        uct_viz.plot_intervals(
            y_pred=y_pred,
            y_std=y_std,
            y_true=y_true,
            n_subset=n_subset,
            ylims=ylims,
            num_stds_confidence_bound=int(num_stds_confidence_bound),
            ax=ax,
        )

    rmse = np.sqrt(np.mean((y_pred - y_true) ** 2))

    ax.text(
        0.05,
        0.95,
        f"RMSE: {rmse:.2f}",
        transform=ax.transAxes,
        verticalalignment="top",
    )


def plot_sharpness(
        y_std: np.ndarray, n_subset: int = 100, ax: Optional[plt.Axes] = None
) -> None:
    """
    Plots sharpness of uncertainty estimates.

    Parameters
    ----------
    y_std : np.ndarray
        Standard deviation of predictions.
    n_subset : int, optional
        Number of samples to plot, by default 100.
    ax : Optional[plt.Axes], optional
        Matplotlib axis to plot on, by default None.

    Returns
    -------
    None
    """
    if ax is None:
        fig, ax = plt.subplots()
    uct_viz.plot_sharpness(y_std=y_std, n_subset=n_subset, ax=ax)


def get_calib_props(
        y_pred: np.ndarray,
        y_std: np.ndarray,
        y_true: np.ndarray,
        vectorized: bool = True,
        prop_type: str = "interval",
        output_folder: Optional[Path] = None,
) -> pd.DataFrame:
    """
    Computes expected and observed proportions for calibration plots.

    Parameters
    ----------
    y_pred : np.ndarray
        Predicted values.
    y_std : np.ndarray
        Standard deviation of predictions.
    y_true : np.ndarray
        True values.
    vectorized : bool, optional
        Whether to use vectorized computations, by default True.
    prop_type : str, optional
        Type of calibration proportion ("interval" or "quantile"), by default "interval".
    output_folder : Optional[Path], optional
        Directory to save calibration data, by default None.

    Returns
    -------
    pd.DataFrame
        DataFrame containing expected and observed proportions.
    """
    if vectorized:
        (
            exp_proportions,
            obs_proportions,
        ) = uct_viz.get_proportion_lists_vectorized(
            y_pred, y_std, y_true, prop_type=prop_type
        )
    else:
        (exp_proportions, obs_proportions) = uct_viz.get_proportion_lists(
            y_pred, y_std, y_true, prop_type=prop_type
        )

    # Create a DataFrame for the calibration plot
    calib_df = pd.DataFrame(
        {
            "Expected Proportion": exp_proportions,
            "Observed Proportion": obs_proportions,
        }
    )

    # Save the calibration plot data
    if output_folder is not None:
        calib_df.to_csv(
            os.path.join(output_folder, "calibration_plot_data.csv"), index=False
        )

    return calib_df


def plot_calibration_from_props():
    raise NotImplementedError("This function is not implemented yet.")


def make_uct_plots(
        y_pred: np.ndarray,
        y_std: np.ndarray,
        y_true: np.ndarray,
        task_name: Optional[str] = None,
        n_subset: int = 100,
        ylims: Optional[Tuple[float, float]] = (-3.0, 3.0),
        num_stds_confidence_bound: float = 2.0,
        plot_save_str: str = "uct_plot",
        savefig: bool = True,
        save_dir: Path = Path("path/to/figures"),
        exp_props: Optional[np.ndarray] = None,
        obs_props: Optional[np.ndarray] = None,
) -> Dict[str, plt.Figure]:
    """
    Generates various Uncertainty Calibration and Analysis plots.

    Parameters
    ----------
    y_pred : np.ndarray
        Model-predicted values.
    y_std : np.ndarray
        Standard deviations of predictions.
    y_true : np.ndarray
        True target values.
    task_name : Optional[str], optional
        Name of the task for labeling, by default None.
    n_subset : int, optional
        Number of samples to include in the plots, by default 100.
    ylims : Optional[Tuple[float, float]], optional
        Limits for the y-axis, by default (-3.0, 3.0).
    num_stds_confidence_bound : float, optional
        Confidence bound in terms of standard deviations, by default 2.0.
    plot_save_str : str, optional
        Prefix for saving plot files, by default "uct_plot".
    savefig : bool, optional
        Whether to save the generated figures, by default True.
    save_dir : Path, optional
        Directory to save the plots, by default "path/to/figures".
    exp_props : Optional[np.ndarray], optional
        Expected proportions for calibration plots, by default None.
    obs_props : Optional[np.ndarray], optional
        Observed proportions for calibration plots, by default None.

    Returns
    -------
    Dict[str, plt.Figure]
        A dictionary containing different generated plots.
    """
    if not all(isinstance(arr, np.ndarray) for arr in [y_pred, y_std, y_true]):
        raise ValueError("y_preds, y_std, and y_true must be numpy arrays.")

    task_name = task_name if task_name is not None else "PCM"
    plots = {}
    plot_functions = [
        (
            "prediction_intervals",
            plot_pred_intervals,
            {
                "ylims": ylims,
                "num_stds_confidence_bound": num_stds_confidence_bound,
                "ordered": False,
            },
        ),
        (
            "ordered_prediction_intervals",
            plot_pred_intervals,
            {
                "ylims": ylims,
                "num_stds_confidence_bound": num_stds_confidence_bound,
                "ordered": True,
            },
        ),
        (
            "calibration",
            uct_viz.plot_calibration,
            {
                "exp_props": exp_props,
                "obs_props": obs_props,
            },
        ),
        (
            "adversarial_group_calibration",
            uct_viz.plot_adversarial_group_calibration,
            {},
        ),
        ("sharpness", plot_sharpness, {}),
        ("residuals_vs_stds", uct_viz.plot_residuals_vs_stds, {}),
        # ("true_vs_predictions", plot_true_vs_preds, {}),
    ]

    for plot_name, plot_func, kwargs in plot_functions:
        fig, ax = plt.subplots(figsize=(6, 4))
        if plot_name == "sharpness":
            plot_func(y_std=y_std, n_subset=n_subset, ax=ax)
        else:
            if "num_stds_confidence_bound" in kwargs:
                kwargs["num_stds_confidence_bound"] = int(kwargs["num_stds_confidence_bound"])  # type: ignore
            plot_func(
                y_pred=y_pred,
                y_std=y_std,
                y_true=y_true,
                n_subset=n_subset,
                ax=ax,
                **kwargs,
            )
        pname = plot_name.replace("_", " ").capitalize()
        t_name = task_name.replace("_", " ")
        ax.set_title(f"{pname} - {t_name}")
        # ax.set_title(f"{plot_name} - {task_name}".replace("_", " ").title())
        plt.tight_layout()
        plots[plot_name] = fig

        if savefig:
            fig_save_path = save_dir / f"{plot_save_str}_{plot_name}"
            uct.viz.save_figure(
                str(fig_save_path),
                ext_list=["png", "svg", "eps"],
                white_background=True,
            )
            # plt.savefig(fig_save_path, format="png", bbox_inches="tight")
    #
    plots["true_vs_predictions"] = plot_true_vs_preds(
        y_pred=y_pred,
        y_true=y_true,
        n_subset=n_subset,
        distant_threshold=3,
        savefig=savefig,
        save_dir=save_dir,
        task_name=task_name,
    )

    return plots


def get_calib_with_recal(
        recal_model: IsotonicRegression,
        y_pred: np.ndarray,
        y_std: np.ndarray,
        y_true: np.ndarray,
        num_bins: int,
        verbose: bool = True,
) -> Dict[str, float]:
    """
    Computes calibration metrics with recalibration applied.

    Parameters
    ----------
    recal_model : IsotonicRegression
        Trained isotonic regression model for recalibration.
    y_pred : np.ndarray
        Predicted values.
    y_std : np.ndarray
        Standard deviations of predictions.
    y_true : np.ndarray
        True target values.
    num_bins : int
        Number of bins for calibration.
    verbose : bool, optional
        Whether to print calibration metrics, by default True.

    Returns
    -------
    Dict[str, float]
        Dictionary containing calibration metrics.
    """
    if verbose:
        print(" (2/n) Calculating average calibration metrics")

    cali_metrics = {}
    cali_metrics["rms_cal"] = uct.root_mean_squared_calibration_error(
        y_pred, y_std, y_true, num_bins=num_bins, recal_model=recal_model
    )
    cali_metrics["ma_cal"] = uct.mean_absolute_calibration_error(
        y_pred, y_std, y_true, num_bins=num_bins, recal_model=recal_model
    )
    cali_metrics["miscal_area"] = uct.miscalibration_area(
        y_pred, y_std, y_true, num_bins=num_bins, recal_model=recal_model
    )
    if verbose:
        print(
            " Average Calibration Metrics including recalibration model".center(60, "=")
        )
        for cali_metric, cali_val in cali_metrics.items():
            print("  {:<37} {:.3f}".format(METRIC_NAMES[cali_metric], cali_val))
    return cali_metrics


def calculate_uct_metrics(
        y_pred: np.ndarray,
        y_std: np.ndarray,
        y_true: np.ndarray,
        n_subset: Optional[int] = None,
        Nbins: int = 100,
        task_name: Optional[str] = None,
        figpath: Path = FIGS_DIR,
        exp_props: Optional[np.ndarray] = None,
        obs_props: Optional[np.ndarray] = None,
        recal_model: Optional[IsotonicRegression] = None,
        verbose: bool = True,
) -> Tuple[Dict[str, Dict[str, float]], Dict[str, plt.Figure]]:
    """
    Computes Uncertainty Calibration Tool (UCT) metrics and generates corresponding plots.

    Parameters
    ----------
    y_pred : np.ndarray
        Predicted values.
    y_std : np.ndarray
        Standard deviations of predictions.
    y_true : np.ndarray
        True target values.
    n_subset : Optional[int], optional
        Number of samples to include in plots, by default None.
    Nbins : int, optional
        Number of bins for calibration, by default 100.
    task_name : Optional[str], optional
        Name of the task for labeling, by default None.
    figpath : Path, optional
        Directory to save figures, by default FIGS_DIR.
    exp_props : Optional[np.ndarray], optional
        Expected proportions for calibration plots, by default None.
    obs_props : Optional[np.ndarray], optional
        Observed proportions for calibration plots, by default None.
    recal_model : Optional[IsotonicRegression], optional
        Recalibration model for uncertainty estimates, by default None.
    verbose : bool, optional
        Whether to print metrics, by default True.

    Returns
    -------
    Tuple[Dict[str, Dict[str, float]], Dict[str, plt.Figure]]:
        Dictionary containing computed UCT metrics and generated plots.
    """

    metrics = get_all_metrics(
        y_pred=y_pred,
        y_std=y_std,
        y_true=y_true,
        num_bins=Nbins,
        resolution=99,
        scaled=True,
        verbose=verbose,
    )
    if recal_model is not None:
        # updating this one to include recalibration
        metrics["avg_calibration"] = get_calib_with_recal(
            recal_model, y_pred, y_std, y_true, Nbins, verbose=verbose
        )

    metrics_filepath = Path(figpath) / f"{task_name}_metrics.pkl"
    save_pickle(metrics, metrics_filepath)

    plots = make_uct_plots(
        y_pred,
        y_std,
        y_true,
        task_name=task_name,
        n_subset=n_subset,  # min(200, len(y_true)),
        ylims=None,
        num_stds_confidence_bound=2.0,
        plot_save_str=f"{task_name}_uct",
        savefig=True,
        save_dir=Path(figpath),
        exp_props=exp_props,
        obs_props=obs_props,
    )
    return metrics, plots


def make_uq_plots(
        ordered_df: pd.DataFrame,
        gaus_pred: Union[np.ndarray, list[float | Any]],
        errors_observed: Union[np.ndarray, list[float | Any]],
        mis_cal: float,
        Nbins: int = 100,
        include_bootstrap: bool = True,
        task_name: str = "PCM",
        figpath: Path = FIGS_DIR,
        logger: logging.Logger = logging.Logger("uqtools"),
) -> Tuple[Dict[str, plt.Figure], float, float, float]:
    """
    Generates plots related to uncertainty quantification and miscalibration.

    Parameters
    ----------
    ordered_df : pd.DataFrame
        Dataframe containing ordered uncertainties and errors.
    gaus_pred : Union[np.ndarray, list[float | Any]]
        Gaussian-predicted calibration values.
    errors_observed : Union[np.ndarray, list[float | Any]]
        Observed error calibration values.
    mis_cal : float
        Miscalibration area value.
    Nbins : int, optional
        Number of bins for discretization, by default 100.
    include_bootstrap : bool, optional
        Whether to include bootstrap-based confidence intervals, by default True.
    task_name : str, optional
        Name of the task, by default "PCM".
    figpath : Path, optional
        Directory to save plots, by default FIGS_DIR.
    logger : logging.Logger, optional
        Logger for debugging and information, by default "uqtools".

    Returns
    -------
    Tuple[Dict[str, plt.Figure], float, float, float]
        Dictionary of generated plots and regression calibration slope, R^2, and intercept.
    """
    # generate error-based calibration plot
    fig, slope, r_sq, intercept = get_slope_metric(
        ordered_df.uq.to_numpy(),
        ordered_df.errors.to_numpy(),
        Nbins=Nbins,
        include_bootstrap=include_bootstrap,
        logger=logger,
    )
    if figpath is not None:
        fig.savefig(Path(figpath) / f"{task_name}_rmv_vs_rmse.png", dpi=1200)
        fig.savefig(Path(figpath) / f"{task_name}_rmv_vs_rmse.svg")
        fig.savefig(Path(figpath) / f"{task_name}_rmv_vs_rmse.eps")

    # Generate Z-score plot and calibration curve
    fig2, _ = plot_Z_scores(ordered_df.errors.to_numpy(), ordered_df.uq.to_numpy())
    if figpath is not None:
        fig2.savefig(Path(figpath) / f"{task_name}_Z_scores.png", dpi=1200)
        fig2.savefig(Path(figpath) / f"{task_name}_Z_scores.svg")
        fig2.savefig(Path(figpath) / f"{task_name}_Z_scores.eps")

    fig3 = plot_calibration_curve(gaus_pred, errors_observed, mis_cal)
    if figpath is not None:
        fig3.savefig(Path(figpath) / f"{task_name}_uq_calibration_curve.png", dpi=1200)
        fig3.savefig(Path(figpath) / f"{task_name}_uq_calibration_curve.svg")
        fig3.savefig(Path(figpath) / f"{task_name}_uq_calibration_curve.eps")
    plots = {"rmv_vs_rmse": fig, "Z_scores": fig2, "calibration_curve": fig3}
    return plots, slope, r_sq, intercept


def calculate_uqtools_metrics(
        uncertainties: np.ndarray,
        errors: np.ndarray,
        Nbins: int = 100,
        include_bootstrap: bool = True,
        task_name: str = "PCM",
        figpath: Path = FIGS_DIR,
        logger: logging.Logger = logging.Logger("uqtools"),
) -> Tuple[Dict[str, float], Dict[str, plt.Figure]]:
    """
    Computes various uncertainty quantification metrics and generates related plots.

    Parameters
    ----------
    uncertainties : np.ndarray
        Model-predicted uncertainty estimates.
    errors : np.ndarray
        Absolute errors in predictions.
    Nbins : int, optional
        Number of bins for calibration, by default 100.
    include_bootstrap : bool, optional
        Whether to include bootstrap-based confidence intervals, by default True.
    task_name : str, optional
        Name of the task, by default "PCM".
    figpath : Path, optional
        Directory to save plots, by default FIGS_DIR.
    logger : logging.Logger, optional
        Logger for debugging and logging results, by default "uqtools".

    Returns
    -------
    Tuple[Dict[str, float], Dict[str, plt.Figure]]
        Dictionary of computed uncertainty metrics and generated plots.
    """
    # Order uncertainties and errors according to uncertainties
    ordered_df = order_sig_and_errors(uncertainties, errors)
    # calculate rho_rank and rho_rank_sim # TODO spearman_rank is already implemented in scipy
    # rho_rank, _ = spearman_rank_corr(np.abs(errors), uncertainties)
    rho_rank, _ = spearmanr(np.abs(errors), uncertainties)
    logger.debug(f"rho_rank = {rho_rank:.2f}")
    exp_rhos_temp = []
    for i in range(1000):
        exp_rho, _ = expected_rho(uncertainties)
        exp_rhos_temp.append(exp_rho)
    rho_rank_sim = np.mean(exp_rhos_temp)
    rho_rank_sim_std = np.std(exp_rhos_temp)
    logger.debug(f"rho_rank_sim = {rho_rank_sim:.2f} +/- {rho_rank_sim_std:.2f}")

    # Calculate the miscalibration area
    gaus_pred, errors_observed = calibration_curve(ordered_df.abs_z.to_numpy())
    mis_cal = calibration_area(errors_observed, gaus_pred)
    logger.debug(f"miscalibration area = {mis_cal:.2f}")

    # Calculate NLL and simulated NLL
    _NLL = NLL(uncertainties, errors)
    logger.debug(f"NLL = {_NLL:.2f}")
    exp_NLL = []
    rng = np.random.default_rng()
    for i in range(1000):
        sim_errors = np.abs(rng.normal(0, uncertainties))

        NLL_sim = NLL(uncertainties, sim_errors)
        exp_NLL.append(NLL_sim)
    NLL_sim = np.mean(exp_NLL)
    NLL_sim_std = np.std(exp_NLL)
    logger.debug(f"NLL_sim = {NLL_sim:.2f} +/- {NLL_sim_std:.2f}")

    plots, slope, r_sq, intercept = make_uq_plots(
        ordered_df,
        gaus_pred,
        errors_observed,
        mis_cal,
        Nbins=Nbins,
        include_bootstrap=include_bootstrap,
        task_name=task_name,
        figpath=figpath,
        logger=logger,
    )

    # Z-metrics
    Z = errors / uncertainties
    Z_var = np.var(Z)
    interval_var = bootstrap((Z,), np.var)
    logger.debug(f"var(Z) = {Z_var:.2f} CI = {interval_var.confidence_interval}")
    Z_mean = np.mean(Z)
    interval_mean = bootstrap((Z,), np.mean)
    logger.debug(f"mean(Z) = {Z_mean:.2f} CI = {interval_mean.confidence_interval}")

    metrics = {
        "rho_rank": rho_rank,
        "rho_rank_sim": rho_rank_sim,
        "rho_rank_sim_std": rho_rank_sim_std,
        "uq_mis_cal": mis_cal,
        "uq_NLL": _NLL,
        "uq_NLL_sim": NLL_sim,
        "uq_NLL_sim_std": NLL_sim_std,
        "Z_var": Z_var,
        "Z_var_CI_low": interval_var.confidence_interval.low,
        "Z_var_CI_high": interval_var.confidence_interval.high,
        "Z_mean": Z_mean,
        "Z_mean_CI_low": interval_mean.confidence_interval.low,
        "Z_mean_CI_high": interval_mean.confidence_interval.high,
        "rmv_rmse_slope": slope,
        "rmv_rmse_r_sq": r_sq,
        "rmv_rmse_intercept": intercept,
    }

    return metrics, plots


def order_sig_and_errors(sigmas: np.ndarray, errors: np.ndarray) -> pd.DataFrame:
    """
    Orders uncertainty values and corresponding errors in ascending order of uncertainty.

    Parameters
    ----------
    sigmas : np.ndarray
        Array of uncertainty values.
    errors : np.ndarray
        Array of corresponding errors.

    Returns
    -------
    pd.DataFrame
        A DataFrame containing ordered uncertainties and errors.
    """
    ordered_df = pd.DataFrame()
    ordered_df["uq"] = sigmas
    ordered_df["errors"] = errors
    ordered_df["abs_z"] = np.abs(ordered_df.errors) / ordered_df.uq
    ordered_df = ordered_df.sort_values(by="uq")
    return ordered_df


def rmse(x: np.ndarray, axis: Optional[int] = None) -> float:
    """
    Computes the Root Mean Squared Error (RMSE).

    Parameters
    ----------
    x : np.ndarray
        Input data array.
    axis : Optional[int], optional
        Axis along which to compute RMSE, by default None.

    Returns
    -------
    float
        Computed RMSE value.
    """
    return np.sqrt((x ** 2).mean())


### SOURCE: UQtools.py https://github.com/jensengroup/UQ_validation_methods/blob/main/UQtools.py ###
def get_bootstrap_intervals(
        errors_ordered: np.ndarray, Nbins: int = 10
) -> Tuple[List[float], List[float]]:
    """
    Computes bootstrap confidence intervals for error calibration.

    Parameters
    ----------
    errors_ordered : np.ndarray
        Ordered error values.
    Nbins : int, optional
        Number of bins for calibration, by default 10.

    Returns
    -------
    Tuple[List[float], List[float]]
        Lower and upper confidence intervals for each bin.
    """
    ci_low = []
    ci_high = []
    N_total = len(errors_ordered)
    N_entries = math.ceil(N_total / Nbins)

    for i in range(0, N_total, N_entries):
        data = errors_ordered[i: i + N_entries]
        res = bootstrap((data,), rmse, vectorized=False)
        ci_low.append(res.confidence_interval[0])
        ci_high.append(res.confidence_interval[1])
    return ci_low, ci_high


def expected_rho(uncertainties: np.ndarray) -> Tuple[float, np.ndarray]:
    """
    Simulates expected errors for each uncertainty and calculates Spearman rank correlation.

    Parameters
    ----------
    uncertainties : np.ndarray
        Array of uncertainty values.

    Returns
    -------
    Tuple[float, np.ndarray]
        Spearman correlation coefficient and simulated errors.
    """

    # sim_errors = []
    rng = np.random.default_rng()
    sim_errors = np.abs(rng.normal(0, uncertainties))
    rho, _ = spearmanr(uncertainties, sim_errors)
    return rho, sim_errors


def NLL(uncertainties: np.ndarray, errors: np.ndarray) -> float:
    """
    Computes the Negative Log-Likelihood (NLL) for a given set of uncertainties and errors.

    Parameters
    ----------
    uncertainties : np.ndarray
        Array of uncertainty estimates.
    errors : np.ndarray
        Array of errors.

    Returns
    -------
    float
        Computed NLL value.
    """
    NLL = 0
    for uncertainty, error in zip(uncertainties, errors):
        temp = math.log(2 * np.pi * uncertainty ** 2) + (error) ** 2 / uncertainty ** 2
        NLL += temp

    NLL = NLL / (2 * len(uncertainties))
    return NLL


def calibration_curve(
        errors_sigma: np.ndarray,
) -> tuple[list[int | Any], list[float | Any]]:
    """
    Computes calibration curves based on error sigmas.

    Parameters
    ----------
    errors_sigma : np.ndarray
        Array of standardized errors.

    Returns
    -------
    Tuple[np.ndarray, np.ndarray]
        Gaussian-predicted calibration and observed error calibration.
    """
    N_errors = len(errors_sigma)
    gaus_pred = []
    errors_observed = []
    for i in np.arange(-10, 0 + 0.01, 0.01):
        gaus_int = 2 * norm(loc=0, scale=1).cdf(i)
        gaus_pred.append(gaus_int)
        observed_errors = (errors_sigma > abs(i)).sum()
        errors_frac = observed_errors / N_errors
        errors_observed.append(errors_frac)

    return gaus_pred, errors_observed


def plot_calibration_curve(
        gaus_pred: np.ndarray, errors_observed: np.ndarray, mis_cal: float
) -> plt.Figure:
    """
    Plots a calibration curve comparing expected and observed error fractions.

    Parameters
    ----------
    gaus_pred : np.ndarray
        Gaussian-predicted calibration values.
    errors_observed : np.ndarray
        Observed error calibration values.
    mis_cal : float
        Miscalibration area value.

    Returns
    -------
    plt.Figure
        Generated calibration plot figure.
    """
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.fill_between(
        gaus_pred,
        gaus_pred,
        errors_observed,
        color="purple",
        alpha=0.4,
        label="miscalibration area = {:0.3f}".format(mis_cal),
    )
    ax.plot(gaus_pred, errors_observed, color="purple", alpha=1)
    ax.plot(np.arange(0, 1, 0.01), np.arange(0, 1, 0.01), linestyle="dashed", color="k")
    ax.set_xlabel("expected fraction of errors", fontsize=14)
    ax.set_ylabel("observed fraction of errors", fontsize=14)
    ax.legend(fontsize=14, loc="lower right")
    return fig


def plot_Z_scores(
        errors: np.ndarray, uncertainties: np.ndarray
) -> Tuple[plt.Figure, plt.Axes]:
    """
    Plots a histogram of standardized Z-scores for uncertainty quantification.

    Parameters
    ----------
    errors : np.ndarray
        Array of observed errors.
    uncertainties : np.ndarray
        Array of uncertainty estimates.

    Returns
    -------
    Tuple[plt.Figure, plt.Axes]
        The generated histogram figure and axes.
    """
    Z_scores = errors / uncertainties
    N_bins = 29
    xmin, xmax = -7, 7
    y, bin_edges = np.histogram(Z_scores, bins=N_bins, range=(xmin, xmax))
    bin_width = bin_edges[1] - bin_edges[0]
    x = 0.5 * (bin_edges[1:] + bin_edges[:-1])
    sy = np.sqrt(y)
    target_values = np.array(
        [len(errors) * bin_width * norm.pdf(x_value) for x_value in x]
    )
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.hist(Z_scores, bins=N_bins, range=(xmin, xmax), color="purple", alpha=0.3)
    ax.errorbar(x, y, sy, fmt=".", color="k")
    ax.plot(
        np.arange(-7, 7, 0.1),
        len(errors) * bin_width * norm.pdf(np.arange(-7, 7, 0.1), 0, 1),
        color="k",
    )
    ax.set_xlabel("error (Z)", fontsize=16)
    ax.set_ylabel("count", fontsize=16)
    ax.set_xlim((-7, 7))
    return fig, ax


def take_closest(myList: List[float], myNumber: float) -> Tuple[float, float, int, int]:
    """
    Finds the two closest values to a given number in a sorted list.

    Parameters
    ----------
    myList : List[float]
        Sorted list of numbers.
    myNumber : float
        The number to find closest values for.

    Returns
    -------
    Tuple[float, float, int, int]
        The two closest values and their corresponding indices.
    """
    pos = bisect_left(myList, myNumber)
    if pos == 0:
        return myList[0], myList[1], 0, 1
    if pos == len(myList):
        return myList[-2], myList[-1], -2, -1
    before = myList[pos - 1]
    after = myList[pos]
    if myNumber < before or myNumber > after:
        print("problem")
    else:
        return before, after, pos - 1, pos


def f_linear_segment(x: float, point_list: List[float], x_list: List[float]) -> float:
    """
    Computes a linear interpolation between points.

    Parameters
    ----------
    x : float
        The x-value for interpolation.
    point_list : List[float]
        List of y-values corresponding to x_list.
    x_list : List[float]
        List of x-values.

    Returns
    -------
    float
        Interpolated y-value.
    """
    x1, x2, x1_idx, x2_idx = take_closest(x_list, x)
    f = point_list[x1_idx] + (x - x1) / (x2 - x1) * (
            point_list[x2_idx] - point_list[x1_idx]
    )
    return f


def area_function(
        x: float, observed_list: List[float], predicted_list: List[float]
) -> float:
    """
    Computes the absolute difference between observed and predicted calibration.

    Parameters
    ----------
    x : float
        x-value for integration.
    observed_list : List[float]
        List of observed calibration values.
    predicted_list : List[float]
        List of predicted calibration values.

    Returns
    -------
    float
        Absolute difference between observed and predicted values at x.
    """
    h = np.abs((f_linear_segment(x, observed_list, predicted_list) - x))
    return h


def calibration_area(
        observed: Union[np.ndarray, list[float]], predicted: Union[np.ndarray, list[float]]
) -> float:
    """
    Computes the calibration miscalibration area using numerical integration.

    Parameters
    ----------
    observed : Union[np.ndarray, list[float]]
        Observed calibration values.
    predicted : Union[np.ndarray, list[float]]
        Predicted calibration values.

    Returns
    -------
    float
        Computed miscalibration area.
    """
    area = 0
    x = min(predicted)
    while x < max(predicted):
        temp, _ = quad(area_function, x, x + 0.001, args=(observed, predicted))
        area += temp
        x += 0.001
    return area


def chi_squared(
        x_values: np.ndarray, x_sigmas: np.ndarray, target_values: np.ndarray
) -> Tuple[float, float]:
    """
    Computes the chi-squared statistic and corresponding probability.

    Parameters
    ----------
    x_values : np.ndarray
        Observed values.
    x_sigmas : np.ndarray
        Standard deviations associated with the observed values.
    target_values : np.ndarray
        Expected (target) values.

    Returns
    -------
    Tuple[float, float]
        Chi-squared statistic and corresponding probability.
    """
    mask = x_values > 0
    chi_value = ((x_values[mask] - target_values[mask]) / x_sigmas[mask]) ** 2
    chi_value = np.sum(chi_value)

    N_free_cs = len(x_values[mask])
    print(N_free_cs)
    chi_prob = ss.chi2.sf(chi_value, N_free_cs)
    return chi_value, chi_prob


def get_slope_metric(
        uq_ordered: np.ndarray,
        errors_ordered: np.ndarray,
        Nbins: int = 10,
        include_bootstrap: bool = True,
        logger: logging.Logger = logging.getLogger("uqtools"),
) -> Tuple[plt.Figure, float, float, float]:
    """
    Computes the slope metric for error-based calibration.

    Parameters
    ----------
    uq_ordered : np.ndarray
        Ordered uncertainties in increasing order.
    errors_ordered : np.ndarray
        Observed errors corresponding to the uncertainties.
    Nbins : int, optional
        Number of bins for computing error calibration (default is 10).
    include_bootstrap : bool, optional
        Whether to include bootstrap confidence intervals (default is True).
    logger : logging.Logger, optional
        Logger instance for debugging information (default is "uqtools").

    Returns
    -------
    Tuple[plt.Figure, float, float, float]
        The generated plot, slope, R-squared value, and intercept of the fitted regression.
    """
    rmvs, rmses, ci_low, ci_high = get_rmvs_and_rmses(
        uq_ordered, errors_ordered, Nbins=Nbins, include_bootstrap=include_bootstrap
    )

    x = np.array(rmvs).reshape((-1, 1))
    y = np.array(rmses)
    model = LinearRegression().fit(x, y)

    # Obtain the coefficient of determination by calling the model with the score() function, then print the coefficient:
    r_sq = model.score(x, y)
    logger.debug(f"R squared:{r_sq}")

    # Print the Intercept:
    intercept = model.intercept_
    logger.debug(f"intercept:{intercept}")

    # Print the Slope:
    slope = model.coef_[0]
    logger.debug(f"slope:{slope}")

    # Predict a Response and print it:
    y_pred = model.predict(x)

    fig, ax = plt.subplots(figsize=(8, 5))
    assymetric_errors = [
        np.array(rmses) - np.array(ci_low),
        np.array(ci_high) - np.array(rmses),
    ]
    ax.errorbar(x, y, yerr=assymetric_errors, fmt="o", linewidth=2)
    ax.plot(
        np.arange(rmvs[0], rmvs[-1], 0.0001),
        np.arange(rmvs[0], rmvs[-1], 0.0001),
        linestyle="dashed",
        color="k",
    )
    ax.plot(
        rmvs,
        y_pred,
        linestyle="dashed",
        color="red",
        label=r"$R^2$ = "
              + "{:0.2f}".format(r_sq)
              + ", slope = {:0.2f}".format(slope)
              + ", intercept = {:0.2f}".format(intercept),
    )

    ax.set_xlabel("RMV", fontsize=14)
    ax.set_ylabel("RMSE", fontsize=14)
    ax.legend(fontsize=14)
    return fig, slope, r_sq, intercept


def get_rmvs_and_rmses(
        uq_ordered: np.ndarray,
        errors_ordered: np.ndarray,
        Nbins: int = 10,
        include_bootstrap: bool = True,
) -> Tuple[List[float], List[float], Optional[List[float]], Optional[List[float]]]:
    """
    Computes RMV (root mean variance) and RMSE (root mean squared error) over bins.

    Parameters
    ----------
    uq_ordered : np.ndarray
        Ordered uncertainty values.
    errors_ordered : np.ndarray
        Corresponding errors.
    Nbins : int, optional
        Number of bins to divide the data (default is 10).
    include_bootstrap : bool, optional
        Whether to compute bootstrap confidence intervals (default is True).

    Returns
    -------
    Tuple[List[float], List[float], Optional[List[float]], Optional[List[float]]]
        Lists of RMV values, RMSE values, and optionally lower/upper confidence intervals.
    """
    N_total = len(uq_ordered)
    N_entries = math.ceil(N_total / Nbins)
    rmvs = [
        np.sqrt((uq_ordered[i: i + N_entries] ** 2).mean())
        for i in range(0, N_total, N_entries)
    ]
    rmses = [
        np.sqrt((errors_ordered[i: i + N_entries] ** 2).mean())
        for i in range(0, N_total, N_entries)
    ]
    if include_bootstrap:
        ci_low, ci_high = get_bootstrap_intervals(errors_ordered, Nbins=Nbins)
    else:
        ci_low, ci_high = None, None
    return rmvs, rmses, ci_low, ci_high


def spearman_r(y_test: np.ndarray, y_pred: np.ndarray, y_var: np.ndarray) -> float:
    """
    Computes the Spearman correlation coefficient between squared errors and uncertainty.

    Parameters
    ----------
    y_test : np.ndarray
        Ground truth values.
    y_pred : np.ndarray
        Predicted values.
    y_var : np.ndarray
        Estimated uncertainty values.

    Returns
    -------
    float
        Spearman correlation coefficient.
    """
    return float(spearmanr((y_test - y_pred) ** 2, y_var)[0])


def reliability_diagram(
        y_true: np.ndarray,
        y_pred: np.ndarray,
        n_bins: int = 10,
        ax: Optional[plt.Axes] = None,
        **kwargs,
) -> plt.Axes:
    """
    Plots a reliability diagram comparing predicted and observed fractions.

    Parameters
    ----------
    y_true : np.ndarray
        True values.
    y_pred : np.ndarray
        Predicted values.
    n_bins : int, optional
        Number of bins for calibration (default is 10).
    ax : Optional[plt.Axes], optional
        Matplotlib axes object for plotting (default is None).

    Returns
    -------
    plt.Axes
        Matplotlib axes object with the reliability diagram.
    """
    if ax is None:
        fig, ax = plt.subplots()

    y_true = np.array(y_true)
    y_pred = np.array(y_pred)

    # Calculate the number of samples in each bin
    bin_sizes = np.histogram(y_pred, bins=n_bins)[0]
    bin_edges = np.histogram(y_pred, bins=n_bins)[1]

    # Calculate the mean predicted value in each bin
    bin_means = [
        y_pred[(y_pred >= bin_edges[i]) & (y_pred < bin_edges[i + 1])].mean()
        for i in range(n_bins)
    ]

    # Calculate the mean true value in each bin
    bin_true_means = [
        y_true[(y_pred >= bin_edges[i]) & (y_pred < bin_edges[i + 1])].mean()
        for i in range(n_bins)
    ]

    # Calculate the fraction of true values in each bin
    bin_fractions = [
        y_true[(y_pred >= bin_edges[i]) & (y_pred < bin_edges[i + 1])].mean()
        for i in range(n_bins)
    ]

    # Plot the reliability diagram
    ax.plot(bin_means, bin_fractions, **kwargs)
    ax.plot([0, 1], [0, 1], "k--", label="Perfectly calibrated")
    ax.set_xlabel("Mean predicted value")
    ax.set_ylabel("Fraction of true values")
    ax.set_title("Reliability Diagram")
    ax.legend()

    return ax


class MetricsTable:
    def __init__(
            self,
            config: Optional[Dict[str, Any]] = None,
            model_type: Optional[str] = None,
            add_plots_to_table: bool = False,
            logger: Optional[logging.Logger] = None,
            project_name: Optional[str] = None,
            run_name: Optional[str] = None,
            verbose: bool = True,
            csv_path: Optional[Union[str, Path]] = None,
    ):
        """
        Initializes a metrics table for tracking model performance.

        Parameters
        ----------
        config : Optional[Dict[str, Any]], optional
            Configuration dictionary containing experiment settings (default: None).
        model_type : Optional[str], optional
            Type of the model being evaluated (default: None).
        add_plots_to_table : bool, optional
            Whether to add plots to the metrics table (default: False).
        logger : Optional[logging.Logger], optional
            Logger instance for debugging (default: None).
        project_name : Optional[str], optional
            Name of the project for logging (default: None).
        run_name : Optional[str], optional
            Name of the run for logging (default: None).
        verbose : bool, optional
            Whether to print verbose logs (default: True).
        csv_path : Optional[Union[str, Path]], optional
            Path to save the CSV file for aggregated metrics (default: None).

        Returns
        -------
        None
        """
        self.config = config
        self.logger = logger or create_logger("MetricsTable")
        self.verbose = verbose
        self.activity = config.get("activity_type", "xc50")
        self.split = config.get("split_type", "scaffold_cluster")
        self.model_type = model_type
        self.desc_prot = config.get("descriptor_protein", None)
        self.desc_chem = config.get("descriptor_chemical", None)
        self.mt = config.get("MT", False)
        self.task_type = config.get("task_type", "regression")
        self.data_specific_path = config.get("data_specific_path", None)
        self.model_name = config.get("model_name", "ensemble")
        self.dropout = config.get("dropout", None)
        self.seed = config.get("seed", 42)
        self.add_plots_to_table = add_plots_to_table
        self.wandb_run_name = run_name or wandb.run.name
        self.wandb_project_name = project_name or wandb.run.project

        cols = []
        if self.model_type:
            cols += ["Model type", "Task"]

        cols += [
            "Activity",
            "Split",
            "desc_prot",
            "desc_chem",
            "dropout",
            "seed",
        ]  # Added Seed to the retrieved columns
        if self.task_type == "regression":
            cols.extend(
                [
                    "R2",
                    "RMSE",
                    "MAE",
                    "MDAE",
                    "MARPD",
                    "PCC",  # Pearson correlation coefficient
                    "RMS Calibration",
                    "MA Calibration",
                    "Miscalibration Area",
                    "Sharpness",
                    "NLL",
                    "CRPS",
                    "Check",
                    "Interval",
                ]
            )
            # UQtools
            cols.extend(
                [
                    "rho_rank",
                    "rho_rank_sim",
                    "rho_rank_sim_std",
                    "uq_mis_cal",
                    "uq_NLL",
                    "uq_NLL_sim",
                    "uq_NLL_sim_std",
                    "Z_var",
                    "Z_var_CI_low",
                    "Z_var_CI_high",
                    "Z_mean",
                    "Z_mean_CI_low",
                    "Z_mean_CI_high",
                    "rmv_rmse_slope",
                    "rmv_rmse_r_sq",
                    "rmv_rmse_intercept",
                ]
            )
            if add_plots_to_table:
                self.plot_cols = [
                    "prediction_intervals",
                    "ordered_prediction_intervals",
                    "calibration",
                    "adversarial_group_calibration",
                    "sharpness",
                    "residuals_vs_pred_std",
                    "true_vs_preds",
                    "rmv_vs_rmse",
                    "Z_scores",
                    "calibration_curve",
                ]
                cols.extend(self.plot_cols)

        elif self.task_type == "classification":
            self.plot_cols = None
            cols.extend(
                [
                    "PR-AUC",
                    "range_logAUC",
                    "Accuracy",
                    "Precision",
                    "Recall",
                    "F1",
                    "PR@K",
                    "RP@K",
                ]
            )
        # if self.aleatoric:
        cols.extend(["aleatoric_uct_mean", "epistemic_uct_mean", "total_uct_mean"])
        self.table = wandb.Table(columns=cols)
        if add_plots_to_table:
            self.plot_table = pd.DataFrame(columns=self.plot_cols)
        if csv_path:
            self.df_path = Path(csv_path)
        else:
            self.df_path = (
                    FIGS_DIR / self.data_specific_path / f"{self.wandb_project_name}.csv"
            )

    def __call__(
            self,
            y_pred: np.ndarray,
            y_std: np.ndarray,
            y_true: np.ndarray,
            y_err: np.ndarray,
            y_eps: Optional[np.ndarray] = None,
            n_subset: Optional[int] = None,
            task_name: Optional[str] = None,
            figpath: Optional[Union[str, Path]] = None,
            exp_props: Optional[np.ndarray] = None,
            obs_props: Optional[np.ndarray] = None,
            recal_model: Optional[IsotonicRegression] = None,
            nll: Optional[float] = None,
    ) -> tuple[dict[str, Any] | dict[str, float], dict[Any, Image]]:
        """
        Computes and logs metrics to the table.

        Parameters
        ----------
        y_pred : np.ndarray
            Predicted values.
        y_std : np.ndarray
            Standard deviation (uncertainty) of the predictions.
        y_true : np.ndarray
            True target values.
        y_err : np.ndarray
            Errors between predicted and true values.
        y_eps : Optional[np.ndarray], optional
            Epistemic uncertainty (default: None).
        n_subset : Optional[int], optional
            Number of samples to use for visualization (default: None).
        task_name : Optional[str], optional
            Name of the task (default: None).
        figpath : Optional[Union[str, Path]], optional
            Path to save the figures (default: None).
        exp_props : Optional[np.ndarray], optional
            Expected proportions for recalibration plot (default: None).
        obs_props : Optional[np.ndarray], optional
            Observed proportions for recalibration plot (default: None).
        recal_model : Optional[IsotonicRegression], optional
            Model used for recalibration (default: None).

        Returns
        -------
        Tuple[Dict[str, Any], Dict[str, plt.Figure]]
            A tuple containing the computed metrics and generated plots.
        """
        # y_pred, y_std, y_true, y_err, y_alea = self.not_nan_filter(y_pred, y_std, y_true, y_err)
        if y_eps is None:
            y_eps = np.zeros_like(y_std)
        metrics, plots = self.calculate_metrics(
            y_pred=y_pred,
            y_std=y_std,
            y_true=y_true,
            y_err=y_err,
            y_eps=y_eps,
            n_subset=n_subset,
            task_name=task_name,
            data_specific_path=self.data_specific_path,
            figpath=figpath,
            exp_props=exp_props,
            obs_props=obs_props,
            recal_model=recal_model,
            nll=nll,
        )

        # self.export_plots(imgs, task_name)
        self.add_data(task_name=task_name, metrics=metrics, plots=plots)

        return metrics, plots

    def calculate_metrics(
            self,
            y_pred: np.ndarray,
            y_std: np.ndarray,
            y_true: np.ndarray,
            y_err: np.ndarray,
            y_eps: Optional[np.ndarray] = None,
            data_specific_path: Optional[Union[str, Path]] = None,
            n_subset: Optional[int] = None,
            task_name: Optional[str] = None,
            figpath: Optional[Union[str, Path]] = None,
            exp_props: Optional[np.ndarray] = None,
            obs_props: Optional[np.ndarray] = None,
            recal_model: Optional[IsotonicRegression] = None,
            nll: Optional[float] = None,
    ) -> tuple[dict[str, Any] | dict[str, float], dict[Any, Image]]:
        """
        Computes various uncertainty quantification (UQ) metrics and plots.

        Parameters
        ----------
        y_pred : np.ndarray
            Predicted values.
        y_std : np.ndarray
            Standard deviation / aleatoric uncertainty.
        y_true : np.ndarray
            True target values.
        y_err : np.ndarray
            Errors between predicted and true values.
        y_eps : Optional[np.ndarray], optional
            Epistemic uncertainty (default: None).
        data_specific_path : Optional[Union[str, Path]], optional
            Path to save the results (default: None).
        n_subset : Optional[int], optional
            Number of samples for visualization (default: None).
        task_name : Optional[str], optional
            Name of the task (default: None).
        figpath : Optional[Union[str, Path]], optional
            Path to save figures (default: None).
        exp_props : Optional[np.ndarray], optional
            Expected proportions for calibration plot (default: None).
        obs_props : Optional[np.ndarray], optional
            Observed proportions for calibration plot (default: None).
        recal_model : Optional[IsotonicRegression], optional
            Model for recalibration (default: None).

        Returns
        -------
        Tuple[Dict[str, Any], Dict[str, plt.Figure]]
            A dictionary of computed metrics and a dictionary of generated plots.
        """
        figures_path = Path(
            FIGS_DIR / data_specific_path / self.model_name
            if not figpath and data_specific_path
            else figpath
        )
        figures_path.mkdir(parents=True, exist_ok=True)
        if self.task_type == "regression":
            uctmetrics, uctplots = calculate_uct_metrics(
                y_pred=y_pred,
                y_std=y_std,
                y_true=y_true,
                n_subset=n_subset,
                Nbins=100,  # 100
                task_name=task_name,
                figpath=figures_path,
                exp_props=exp_props,
                obs_props=obs_props,
                recal_model=recal_model,
                verbose=self.verbose,
            )
            if nll:
                uctmetrics["scoring_rule"]["nll"] = nll

            # calculate other uqtools metrics
            uqmetrics, uqplots = calculate_uqtools_metrics(
                y_std + y_eps,
                y_err,
                Nbins=100,  # 100
                include_bootstrap=True,
                task_name=task_name,
                figpath=figures_path,
                logger=self.logger,
            )
            metrics = {**uctmetrics, **uqmetrics}
            plots = {**uctplots, **uqplots}

        else:
            # There is no UQ metrics for classification implemented yet
            metrics = {}
            plots = {}

        if y_eps is None:
            (
                metrics["aleatoric_uct_mean"],
                metrics["epistemic_uct_mean"],
                metrics["total_uct_mean"],
            ) = (None, None, None)
        else:
            # y_alea_mean = y_alea.mean()
            y_eps_mean = y_eps.mean()
            y_alea_mean = y_std.mean()
            # y_std_mean = y_std.mean()
            metrics["aleatoric_uct_mean"] = y_alea_mean
            metrics["epistemic_uct_mean"] = y_eps_mean
            metrics["total_uct_mean"] = y_alea_mean + y_eps_mean

        plots = {k: wandb.Image(v) for k, v in plots.items()}

        return metrics, plots

    def add_data(
            self,
            task_name: str,
            metrics: Dict[str, Any],
            plots: Dict[str, plt.Figure] | Dict[str, Image],
    ) -> None:
        """
        Adds computed metrics and plots to the metrics table.

        Parameters
        ----------
        task_name : str
            Name of the task.
        metrics : Dict[str, Any]
            Dictionary containing computed metrics.
        plots : Dict[str, plt.Figure] or Dict[str, Image]
            Dictionary containing generated plots.

        Returns
        -------
        None
        """
        vals = [self.model_type] if self.model_type is not None else []
        vals.append(task_name)
        vals.extend(
            [
                self.activity,
                self.split,
                self.desc_prot,
                self.desc_chem,
                self.dropout,
                self.seed,
            ]
        )
        if self.task_type == "regression":
            vals.extend(
                [
                    metrics["accuracy"]["r2"],
                    metrics["accuracy"]["rmse"],
                    metrics["accuracy"]["mae"],
                    metrics["accuracy"]["mdae"],
                    metrics["accuracy"]["marpd"],
                    metrics["accuracy"]["corr"],
                    metrics["avg_calibration"]["rms_cal"],
                    metrics["avg_calibration"]["ma_cal"],
                    metrics["avg_calibration"]["miscal_area"],
                    metrics["sharpness"]["sharp"],
                    metrics["scoring_rule"]["nll"],
                    metrics["scoring_rule"]["crps"],
                    metrics["scoring_rule"]["check"],
                    metrics["scoring_rule"]["interval"],
                ]
            )
            # UQtools
            vals.extend(
                [
                    metrics["rho_rank"],
                    metrics["rho_rank_sim"],
                    metrics["rho_rank_sim_std"],
                    metrics["uq_mis_cal"],
                    metrics["uq_NLL"],
                    metrics["uq_NLL_sim"],
                    metrics["uq_NLL_sim_std"],
                    metrics["Z_var"],
                    metrics["Z_var_CI_low"],
                    metrics["Z_var_CI_high"],
                    metrics["Z_mean"],
                    metrics["Z_mean_CI_low"],
                    metrics["Z_mean_CI_high"],
                    metrics["rmv_rmse_slope"],
                    metrics["rmv_rmse_r_sq"],
                    metrics["rmv_rmse_intercept"],
                ]
            )
            if self.add_plots_to_table:
                # plots
                vals.extend(
                    [
                        plots["prediction_intervals"],
                        plots["ordered_prediction_intervals"],
                        plots["calibration"],
                        plots["adversarial_group_calibration"],
                        plots["sharpness"],
                        plots["residuals_vs_stds"],
                        plots["true_vs_predictions"],
                        # UQ tools
                        plots["rmv_vs_rmse"],
                        plots["Z_scores"],
                        plots["calibration_curve"],
                    ]
                )
        elif self.task_type == "classification":
            vals.extend(
                [
                    metrics["PR-AUC"],
                    metrics["range_logAUC"],
                    metrics["Accuracy"],
                    metrics["Precision"],
                    metrics["Recall"],
                    metrics["F1"],
                    metrics["PR@K"],
                    metrics["RP@K"],
                ]
            )
        # if self.aleatoric:
        vals.extend(
            [
                metrics["aleatoric_uct_mean"],
                metrics["epistemic_uct_mean"],
                metrics["total_uct_mean"],
            ]
        )

        self.table.add_data(*vals)
        plt.close()

    def excel_log(self, df: pd.DataFrame, path: Path, add_plots: bool = False) -> None:
        """
        Saves the computed metrics to an Excel file.

        Parameters
        ----------
        df : pd.DataFrame
            DataFrame containing metrics.
        path : Path
            Path to save the Excel file.
        add_plots : bool, optional
            Whether to include plots in the Excel file (default: False).

        Returns
        -------
        None
        """
        if add_plots:
            raise NotImplementedError
        else:
            # dropping the wandb objects
            df.drop(columns=self.plot_cols, inplace=True, errors="ignore")
        df["wandb project"] = self.wandb_project_name
        df["wandb run"] = self.wandb_run_name

        if path.exists():
            book = load_workbook(path)
            writer = pd.ExcelWriter(path, engine="openpyxl")
            writer.book = book
            startrow = writer.sheets["Sheet1"].max_row
        else:
            writer = pd.ExcelWriter(path, engine="openpyxl")
            startrow = 0

        df.to_excel(
            writer,
            sheet_name="Sheet1",
            startrow=startrow,
            header=startrow == 0,
            index=False,
        )
        writer.close()

    def csv_log(self) -> None:
        """
        Saves the computed metrics to a CSV file.

        Returns
        -------
        None
        """
        df = self.table.get_dataframe()
        df["wandb project"] = self.wandb_project_name
        df["wandb run"] = self.wandb_run_name
        df["model name"] = self.model_name

        df.drop(columns=self.plot_cols, inplace=True, errors="ignore")
        df.to_csv(
            self.df_path,
            index=False,
            mode="a" if self.df_path.exists() else "w",
            header=not self.df_path.exists(),
        )

    def wandb_log(self) -> None:
        """
        Logs the metrics table to Weights & Biases (wandb).

        Returns
        -------
        None
        """
        self.csv_log()
        wandb.log(
            {
                f"Uncertainty Metrics Table - {self.model_type} - {wandb.run.name}": self.table
            }
        )


def isotonic_recalibrator(
        y_true_recal: np.ndarray,
        y_pred_recal: np.ndarray,
        y_std_recal: np.ndarray,
        y_true_test: np.ndarray,
        y_pred_test: np.ndarray,
        y_std_test: np.ndarray,
) -> Tuple[IsotonicRegression, np.ndarray, np.ndarray]:
    """
    Trains an isotonic regression model for recalibration and applies it to test predictions.

    Parameters
    ----------
    y_true_recal : np.ndarray
        True target values for recalibration.
    y_pred_recal : np.ndarray
        Predicted values for recalibration.
    y_std_recal : np.ndarray
        Predicted uncertainties for recalibration.
    y_true_test : np.ndarray
        True target values for testing.
    y_pred_test : np.ndarray
        Predicted values for testing.
    y_std_test : np.ndarray
        Predicted uncertainties for testing.

    Returns
    -------
    Tuple[IsotonicRegression, np.ndarray, np.ndarray]
        Recalibration model, expected proportions, and observed proportions after recalibration.
    """

    y_pred_recal = y_pred_recal.flatten()
    y_std_recal = y_std_recal.flatten()
    exp_props, obs_props = uct.metrics_calibration.get_proportion_lists_vectorized(
        y_pred_recal,
        y_std_recal,
        y_true_recal,
    )
    # Train a recalibration model.
    recal_model = uct.recalibration.iso_recal(exp_props, obs_props)
    # Get the expected props and observed props using the new recalibrated model
    te_recal_exp_props, te_recal_obs_props = (
        uct.metrics_calibration.get_proportion_lists_vectorized(
            y_pred_test, y_std_test, y_true_test, recal_model=recal_model
        )
    )

    return recal_model, te_recal_exp_props, te_recal_obs_props


def std_recalibrator(
        y_true_recal: np.ndarray,
        y_pred_recal: np.ndarray,
        y_std_recal: np.ndarray,
        y_std_test: np.ndarray,
) -> Tuple[Callable[[np.ndarray], np.ndarray], np.ndarray, np.ndarray]:
    """
    Recalibrates standard deviation predictions using a recalibration model.

    Parameters
    ----------
    y_true_recal : np.ndarray
        True target values for recalibration.
    y_pred_recal : np.ndarray
        Predicted values for recalibration.
    y_std_recal : np.ndarray
        Predicted uncertainties for recalibration.
    y_std_test : np.ndarray
        Predicted uncertainties for testing.

    Returns
    -------
    Tuple[Callable[[np.ndarray], np.ndarray], np.ndarray, np.ndarray]
        Recalibration function, recalibrated standard deviations for recalibration and testing.
    """
    std_recal = uct.recalibration.get_std_recalibrator(
        y_pred_recal, y_std_recal, y_true_recal
    )

    y_std_recal_recalibrated = std_recal(y_std_recal)
    y_std_test_recalibrated = std_recal(y_std_test)

    return std_recal, y_std_recal_recalibrated, y_std_test_recalibrated


def recalibration_metrics_and_plots(
        y_pred_test: np.ndarray,
        y_std_test: np.ndarray,
        y_true_test: np.ndarray,
        y_err_test: np.ndarray,
        uct_logger: Optional[Callable] = None,
        exp_props: Optional[np.ndarray] = None,
        obs_props: Optional[np.ndarray] = None,
        recal_model: Optional[IsotonicRegression] = None,
        n_subset: Optional[int] = None,
        task_name: str = "before_calibration",
        figpath: Optional[Union[str, Path]] = None,
) -> Tuple[Dict[str, Any], Dict[str, plt.Figure]]:
    """
    Computes recalibration metrics and generates plots.

    Parameters
    ----------
    y_pred_test : np.ndarray
        Predicted values for testing.
    y_std_test : np.ndarray
        Predicted uncertainties for testing.
    y_true_test : np.ndarray
        True target values for testing.
    y_err_test : np.ndarray
        Prediction errors.
    uct_logger : Optional[Callable], optional
        Logger function for uncertainty calibration tracking (default: None).
    exp_props : Optional[np.ndarray], optional
        Expected proportions for calibration plot (default: None).
    obs_props : Optional[np.ndarray], optional
        Observed proportions for calibration plot (default: None).
    recal_model : Optional[IsotonicRegression], optional
        Recalibration model (default: None).
    n_subset : Optional[int], optional
        Number of samples to use for visualization (default: None).
    task_name : str, optional
        Name of the task (default: "before_calibration").
    figpath : Optional[Union[str, Path]], optional
        Path to save the figures (default: None).

    Returns
    -------
    Tuple[Dict[str, Any], Dict[str, plt.Figure]]
        Computed metrics and generated plots.
    """
    if uct_logger is not None:
        metrics, plots = uct_logger(
            y_pred=y_pred_test,
            y_std=y_std_test,
            y_true=y_true_test,
            y_err=y_err_test,
            y_eps=None,
            n_subset=n_subset,
            task_name=task_name,
            figpath=figpath,
            exp_props=exp_props,
            obs_props=obs_props,
            recal_model=recal_model,
        )

    else:  # without logging to wandb then
        uctmetrics, uctplots = calculate_uct_metrics(
            y_pred=y_pred_test,
            y_std=y_std_test,
            y_true=y_true_test,
            n_subset=n_subset,
            Nbins=100,
            task_name=task_name,
            figpath=figpath,
            exp_props=exp_props,
            obs_props=obs_props,
            recal_model=recal_model,
        )
        uqmetrics, uqplots = calculate_uqtools_metrics(
            y_std_test,
            y_err_test,
            Nbins=n_subset,
            include_bootstrap=True,
            task_name=task_name,
            figpath=figpath,
        )
        metrics, plots = {**uctmetrics, **uqmetrics}, {
            **uctplots,
            **uqplots,
        }

    plt.close()

    return metrics, plots


def recalibrate(
        y_true_recal: np.ndarray,
        y_pred_recal: np.ndarray,
        y_alea_recal: np.ndarray,
        y_err_recal: np.ndarray,
        y_true_test: np.ndarray,
        y_pred_test: np.ndarray,
        y_alea_test: np.ndarray,
        y_err_test: np.ndarray,
        n_subset: Optional[int] = None,
        task_name: str = "PCM",
        savefig: bool = True,
        save_dir: Path = Path("path/to/figures"),
        uct_logger: Optional[Callable] = None,
) -> IsotonicRegression:
    """
    Performs uncertainty recalibration using isotonic regression and saves recalibration results.

    Parameters
    ----------
    y_true_recal : np.ndarray
        True target values for recalibration.
    y_pred_recal : np.ndarray
        Predicted values for recalibration.
    y_alea_recal : np.ndarray
        Aleatoric uncertainty for recalibration.
    y_err_recal : np.ndarray
        Prediction errors for recalibration.
    y_true_test : np.ndarray
        True target values for testing.
    y_pred_test : np.ndarray
        Predicted values for testing.
    y_alea_test : np.ndarray
        Aleatoric uncertainty for testing.
    y_err_test : np.ndarray
        Prediction errors for testing.
    n_subset : Optional[int], optional
        Number of samples for visualization (default: None).
    task_name : str, optional
        Name of the task (default: "PCM").
    savefig : bool, optional
        Whether to save the recalibration plots (default: True).
    save_dir : Path, optional
        Path to save the figures and recalibration results (default: "path/to/figures").
    uct_logger : Optional[Callable], optional
        Logger function for uncertainty calibration tracking (default: None).

    Returns
    -------
    IsotonicRegression
        Trained isotonic regression model for recalibration.
    """
    if savefig:
        before_path = Path(save_dir) / "Before_recal"
        after_path = Path(save_dir) / "After_recal"
        before_path.mkdir(exist_ok=True)
        after_path.mkdir(exist_ok=True)

    else:
        before_path = None
        after_path = None
    # Before Calibration
    before_metrics, before_plots = recalibration_metrics_and_plots(
        y_pred_test=y_pred_test,
        y_std_test=y_alea_test,
        y_true_test=y_true_test,
        y_err_test=y_err_test,
        uct_logger=uct_logger,
        n_subset=n_subset,
        task_name=task_name + "_before_calibration",
        figpath=before_path,
    )
    # subset
    subbefore_metrics, subbefore_plots = recalibration_metrics_and_plots(
        y_pred_test=y_pred_test,
        y_std_test=y_alea_test,
        y_true_test=y_true_test,
        y_err_test=y_err_test,
        uct_logger=uct_logger,
        n_subset=100,
        task_name=task_name + "_before_calibration_subset_100",
        figpath=before_path,
    )

    # Recalibrating
    # * this is the isotonic regression one *
    iso_recal_model, te_recal_exp_props, te_recal_obs_props = isotonic_recalibrator(
        y_true_recal=y_true_recal,
        y_pred_recal=y_pred_recal,
        y_std_recal=y_alea_recal,
        y_true_test=y_true_test,
        y_pred_test=y_pred_test,
        y_std_test=y_alea_test,
    )

    # AFter Isotonic metrics
    recal_metrics, recal_plots = recalibration_metrics_and_plots(
        y_pred_test=y_pred_test,
        y_std_test=y_alea_test,
        y_true_test=y_true_test,
        y_err_test=y_err_test,
        uct_logger=uct_logger,
        exp_props=te_recal_exp_props,
        obs_props=te_recal_obs_props,
        recal_model=iso_recal_model,
        n_subset=n_subset,
        task_name=task_name + "_after_calibration_with_isotonic_regression",
        figpath=after_path,
    )
    subrecal_metrics, subrecal_plots = recalibration_metrics_and_plots(
        y_pred_test=y_pred_test,
        y_std_test=y_alea_test,
        y_true_test=y_true_test,
        y_err_test=y_err_test,
        uct_logger=uct_logger,
        exp_props=te_recal_exp_props,
        obs_props=te_recal_obs_props,
        recal_model=iso_recal_model,
        n_subset=100,
        task_name=task_name + "_after_calibration_with_isotonic_regression_subset_100",
        figpath=after_path,
    )

    # Let's save the recalibration model
    recal_model_path = Path(save_dir) / "iso_recalibration_model.pkl"
    save_pickle(iso_recal_model, recal_model_path)

    # recal_model_path = Path(save_dir) / "std_recalibration_model.pkl"
    # save_pickle(std_recal, recal_model_path)

    # save metrics after recalibration with isotonic regression
    recal_metrics_path = Path(save_dir) / "recal_metrics_iso.pkl"
    save_pickle(recal_metrics, recal_metrics_path)
    # TODO : Add evidential NLL as an argument here.
    return iso_recal_model  # std_recal  # , metrics, plots


def calc_alea_epi_mean_var_notnan(
        vars_all: torch.Tensor, targets_all: torch.Tensor
) -> Tuple[torch.Tensor, torch.Tensor]:
    """
    Calculates the mean and variance of aleatoric and epistemic uncertainty,
    considering only valid (non-NaN) corresponding targets.

    Parameters
    ----------
    vars_all : torch.Tensor
        Aleatoric variances from the model, shape [batch_size, num_tasks, num_models].
    targets_all : torch.Tensor
        True target values, shape [batch_size, num_tasks].

    Returns
    -------
    Tuple[torch.Tensor, torch.Tensor]
        Mean and variance of vars_all considering only valid entries.
    """
    # Create a mask of valid (non-NaN) targets
    valid_mask = ~torch.isnan(targets_all)

    # Flatten the valid_vars to a 1D array for mean and variance calculation
    valid_vars_flat = vars_all[valid_mask]

    # Calculate mean and variance on the flattened valid data
    vars_mean = torch.mean(valid_vars_flat.detach())
    vars_var = torch.var(valid_vars_flat.detach())

    return vars_mean, vars_var


def aggregate_metrics_csv(input_file_path: str, output_file_path: str) -> None:
    """
    Aggregates a CSV file by specified columns, calculating mean and standard deviation
    for numeric columns and collecting string columns into lists.

    Parameters
    ----------
    input_file_path : str
        Path to the input CSV file.
    output_file_path : str
        Path to save the aggregated CSV file.

    Returns
    -------
    None

    Example
    -------
    >>> input_file_path = '/path/to/your/inputfile.csv'
    >>> output_file_path = '/path/to/your/outputfile.csv'
    >>> aggregate_metrics_csv(input_file_path, output_file_path)
    """
    # Step 1: Read the CSV file into a DataFrame
    df = pd.read_csv(input_file_path, header=0)

    # Step 2: Define the columns for grouping and the columns for aggregation
    group_cols = ["Model type", "Task", "Activity", "Split", "desc_prot", "desc_chem"]
    numeric_cols = [
        "RMSE",
        "R2",
        "MAE",
        "MDAE",
        "MARPD",
        "PCC",
        "RMS Calibration",
        "MA Calibration",
        "Miscalibration Area",
        "Sharpness",
        "NLL",
        "CRPS",
        "Check",
        "Interval",
        "rho_rank",
        "rho_rank_sim",
        "rho_rank_sim_std",
        "uq_mis_cal",
        "uq_NLL",
        "uq_NLL_sim",
        "uq_NLL_sim_std",
        "Z_var",
        "Z_var_CI_low",
        "Z_var_CI_high",
        "Z_mean",
        "Z_mean_CI_low",
        "Z_mean_CI_high",
        "rmv_rmse_slope",
        "rmv_rmse_r_sq",
        "rmv_rmse_intercept",
        "aleatoric_uct_mean",
        "epistemic_uct_mean",
        "total_uct_mean",
    ]
    string_cols = ["wandb project", "wandb run", "model name"]

    # Step 3: Group the DataFrame by the specified columns
    grouped = df.groupby(group_cols)

    # Step 4: Aggregate the numeric columns
    aggregated = grouped[numeric_cols].agg(["mean", "std"])

    # Combine mean and std into the required format
    for col in numeric_cols:
        aggregated[(col, "combined")] = (
                aggregated[(col, "mean")].round(3).astype(str)
                + " ("
                + aggregated[(col, "std")].round(3).astype(str)
                + ")"
        )

    # Drop the separate mean and std columns, keeping only the combined column
    aggregated = aggregated[[col for col in aggregated.columns if col[1] == "combined"]]

    # Rename the columns to a simpler format
    aggregated.columns = [col[0] for col in aggregated.columns]

    # Step 5: Aggregate the string columns into lists
    string_aggregated = grouped[string_cols].agg(lambda x: list(x))

    # Combine the numeric and string aggregations
    final_aggregated = pd.concat([aggregated, string_aggregated], axis=1).reset_index()

    # Step 6: Write the aggregated results into a new CSV file
    final_aggregated.to_csv(output_file_path, index=False)

    print("Aggregation complete and results saved to", output_file_path)
